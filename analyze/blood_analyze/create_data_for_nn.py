"""
nn_data.xlsxを作成するプログラム

実行条件 
analyze_setting.txtへのデータディレクトリの記述
summary.xlsxのanalyze_infoシートの記入とdataシートの生成が済んでいること
"""

import pandas as pd
import math
from scipy.interpolate import Akima1DInterpolator
import numpy as np
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
import os
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows

from create_analyze_info_check import (
    get_setting_file_path,
    get_data_dir_path,
    read_dir_name_from_settings,
)

DEVIATION_IN_ONE_CYCLE = 0.014285714
CIRCLE_PERIOD = 3
CIRCLE_HIDE = 2.7

ADD_RANGE = 3
POINT = 1000

WINDOW_SIZE = 3

OUTPUT_DATA_CYCLE = 0.15


def read_directory_path_from_settings(file_name):
    """
    Analyze_setting.txtからデータディレクトリのパスを読み取ります。

    Args:
        file_name (str): 設定ファイルの名前。

    Returns:
        str: データディレクトリのパス。ファイルのフォーマットが正しくない場合はNoneを返します。
    """
    with open(file_name, "r", encoding="utf-8") as file:
        lines = file.readlines()
        for i, line in enumerate(lines):
            line = line.strip()
            if line == "directory_path" and i + 1 < len(lines):
                return lines[i + 1].strip()
    return None


def get_deviation(analyze_info):
    """
    Analyze_infoデータフレームからズレ(deviation)のリストを取得します。

    Args:
        analyze_info (DataFrame): 分析情報が含まれるPandasのデータフレーム。

    Returns:
        List[float]: 分析情報から読み取ったズレのリスト。
    """
    deviation = (
        analyze_info[analyze_info[0] == "deviation"].iloc[0, 1:].dropna().tolist()
    )

    return deviation


def get_target_numbers(analyze_info):
    """
    Analyze_infoデータフレームから目標値(target number)のリストを取得します。

    Args:
        analyze_info (DataFrame): 分析情報が含まれるPandasのデータフレーム。

    Returns:
        List[List[float]]: 目標値のリストのリスト。各サブリストは一つの実験データに対応します。
    """
    target_rows = analyze_info[analyze_info[0].str.contains("target number")]
    target_numbers = target_rows.iloc[:, 1:].values.tolist()

    for i in range(len(target_numbers)):
        target_numbers[i] = [num for num in target_numbers[i]]

    target_numbers = [list(row) for row in zip(*target_numbers)]

    target_numbers = [
        [x for x in sublist if not math.isnan(x)] for sublist in target_numbers
    ]

    return target_numbers


def load_sheets(file_path):
    """
    指定されたエクセルファイルから必要なシートを読み込み、データフレームの辞書として返します。

    Args:
        file_path (str): エクセルファイルのパス。

    Returns:
        dict: キーがシート名、値が対応するデータフレームの辞書。
    """
    # エクセルファイルを読み込む
    excel_file = pd.ExcelFile(file_path)

    # エクセルファイルのすべてのシート名を取得する
    sheet_names = excel_file.sheet_names

    # ロードしたくないシートをフィルタリングする
    # analyze_info'を無視し、CH7からCH16のシートをロードしたい。
    filtered_sheets = [
        name
        for name in sheet_names
        if name not in ["analyze_info"] and name.startswith("CH")
    ]

    # 必要なシートをデータフレームの辞書に読み込む
    dataframes = {}
    for sheet_name in filtered_sheets:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        df.columns = [
            f"data{i+1}" for i in range(len(df.columns))
        ]  # 列名を指定(data1,data2,...)
        dataframes[sheet_name] = df

    return dataframes


def apply_moving_average(dataframes, window_size=WINDOW_SIZE):
    """
    各データフレーム内の全ての列に移動平均を適用します。

    Args:
        dataframes (dict): データフレームの辞書。
        window_size (int): 移動平均ウィンドウのサイズ。

    Returns:
        dict: 各列に移動平均を適用したデータフレームの辞書。
    """
    ma_dataframes = {}
    for sheet_name, df in dataframes.items():
        ma_df = df.rolling(window=window_size, min_periods=1).mean()
        ma_dataframes[sheet_name] = ma_df

    return ma_dataframes


def apply_akima_interpolation(dataframes):
    """
    各データフレームの全ての列に秋間補間を適用します。

    Args:
        dataframes (dict): データフレームの辞書。

    Returns:
        dict: 各サブ辞書には、データフレームの各列に対する秋間補間オブジェクトが含まれる辞書。
    """
    akima_interpolators = {}
    for sheet_name, df in dataframes.items():
        interpolators = {}
        for column in df.columns:
            # NaN値を持つ列の処理
            mask = ~np.isnan(df[column])
            if mask.sum() < 2:
                # 補間するのに十分なデータポイントがない
                interpolator = None
            else:
                # カラム用の秋間インターポレーターを作成する。
                data_length = df[column][mask].count()
                x_axis = [round((i * 0.2), 1) for i in range(data_length.astype(int))]
                # print(x_axis) DEBUG
                interpolator = Akima1DInterpolator(x_axis, df[column][mask])
            interpolators[column] = interpolator
        akima_interpolators[sheet_name] = interpolators

    return akima_interpolators


def plot_series(series):
    """
    DEBUG用
    pandas.Seriesのデータをプロットします。

    Args:
        series (pandas.Series): プロットするpandas.Seriesオブジェクト。
    """
    plt.figure(figsize=(10, 6))
    plt.plot(series)
    plt.xlabel("Index")
    plt.ylabel("Value")
    plt.title(f"Line plot of {series.name}")
    plt.show()


def plot_akima_interpolation(akima_interp):
    """
    DEBUG用
    秋間補間オブジェクトを受け取り、それを図示します。

    Args:
        akima_interp: 秋間補間オブジェクト。
    """
    # 補間用の細かいポイントを生成
    x_new = np.linspace(0, 129, 500)
    y_new = akima_interp(x_new)

    # 元のデータポイントと補間曲線をプロット
    plt.figure(figsize=(8, 6))
    plt.plot(x_new, y_new, label="Akima interpolation")
    plt.legend()
    plt.show()


def target_number_to_time(deviation, target_numbers):
    """
    目標値(target number)とズレ(deviation)を時間に変換します。

    Args:
        deviation (List[float]): ズレのリスト。
        target_numbers (List[List[float]]): 目標値のリスト。

    Returns:
        List[List[float]]: ターゲットの提示時間を表す二次元リスト。
    """
    target_time = [
        [target_number * CIRCLE_PERIOD + CIRCLE_HIDE for target_number in sublist]
        for sublist in target_numbers
    ]

    target_time = [
        [
            target_number + deviation_item * DEVIATION_IN_ONE_CYCLE
            for target_number in target_time_sublist
        ]
        for deviation_item, target_time_sublist in zip(deviation, target_time)
    ]
    # DEBUG
    # print(target_time)
    # 正しく時間に変換されていました！

    return target_time


def transform_dicdic(dicdic):
    new_dicdic = {}

    for outer_key, inner_dict in dicdic.items():
        for inner_key, value in inner_dict.items():
            if inner_key not in new_dicdic:
                new_dicdic[inner_key] = {}
            new_dicdic[inner_key][outer_key] = value

    return new_dicdic


def make_target_flag(target_time, x_time):
    target_flag = [0] * len(x_time)

    for i, t in enumerate(x_time):
        for target_value in target_time:
            if target_value <= t <= target_value + CIRCLE_PERIOD:
                target_flag[i] = 1
                break

    return target_flag


def make_data_df(target_time, data_length, ch_interpolators):
    all_data = {}
    keys_ordered = list(ch_interpolators.keys())

    x_time = [
        round((i * OUTPUT_DATA_CYCLE), 2)
        for i in range(int(data_length / OUTPUT_DATA_CYCLE))
    ]

    all_data["Time"] = x_time
    for i in range(len(ch_interpolators)):
        all_data[keys_ordered[i]] = ch_interpolators[keys_ordered[i]](x_time)

    target_flag = make_target_flag(target_time, x_time)
    all_data["target_flag"] = target_flag

    # 最長の列の長さを見つける
    max_length = max(len(lst) for lst in all_data.values())

    for key in all_data.keys():
        length_diff = max_length - len(all_data[key])
        if length_diff > 0:
            all_data[key] = np.append(all_data[key], np.full(length_diff, np.nan))

    all_data_df = pd.DataFrame(all_data)
    all_data_df = all_data_df[:-1]

    return all_data_df


def output_excel_df(nn_data_path, sheet_name, df):
    df.fillna("", inplace=True)
    # Excelファイルの存在確認と開く
    if os.path.exists(nn_data_path):
        book = openpyxl.load_workbook(nn_data_path)
    else:
        book = openpyxl.Workbook()

    sheet_name = sheet_name[:30]

    # シートの処理
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)
    sheet = book.create_sheet(sheet_name)

    # データフレームの書き込みとフォントの適用
    yu_gothic_font = Font(name="Yu Gothic")
    df.fillna("")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.font = yu_gothic_font

    # ファイルの保存前にデフォルトシートを削除
    if "Sheet" in book.sheetnames and len(book.sheetnames) > 1:
        std = book["Sheet"]
        book.remove(std)

    # ファイルの保存
    book.save(nn_data_path)


def main():
    """
    プログラムのメイン関数。設定ファイルの読み込み、データ処理、結果のExcel出力を行います。
    """
    setting_file = get_setting_file_path()
    data_dir_name = read_dir_name_from_settings(setting_file)
    data_dir_path = get_data_dir_path(data_dir_name)

    summary_path = os.path.join(data_dir_path, "blood_excel", "summary.xlsx")
    nn_data_path = os.path.join(data_dir_path, "blood_excel", "nn_data.xlsx")

    if not os.path.isfile(summary_path):
        print("blood_excelディレクトリにsummary.xlsxがないです")
        return

    analyze_info = pd.read_excel(
        summary_path, sheet_name="analyze_info", header=None, engine="openpyxl"
    )

    deviation = get_deviation(analyze_info)  # 1次元リスト

    target_numbers = get_target_numbers(analyze_info)  # 2次元リスト

    data_dic = load_sheets(summary_path)

    ma_data_dic = apply_moving_average(data_dic)

    # DEBUG
    # plot_series(ma_data_dic["CH7"]["data1"])
    # 移動平均は上手く行えていました！

    ch_interpolators_dicdic = apply_akima_interpolation(ma_data_dic)

    # DEBUG
    # plot_akima_interpolation(ch_interpolators["CH7"]["data1"])
    # 補間はうまく行えていました！

    if (
        len(deviation) != len(target_numbers)
        or len(target_numbers) != len(ch_interpolators_dicdic["CH7"])
        or len(deviation) != len(ch_interpolators_dicdic["CH7"])
    ):
        print("ファイル数やanalyze_infoの内容がおかしい")
        return

    target_time = target_number_to_time(deviation, target_numbers)
    data_lengths = ma_data_dic["CH7"].count().tolist()
    data_lengths = [round(data_length * 0.2, 1) for data_length in data_lengths]
    data_interpolators_dicdic = transform_dicdic(ch_interpolators_dicdic)

    i = 0
    for data_number, ch_interpolators in data_interpolators_dicdic.items():
        data_df = make_data_df(target_time[i], data_lengths[i], ch_interpolators)
        output_excel_df(nn_data_path, data_number + "_" + data_dir_name, data_df)
        i += 1


if __name__ == "__main__":
    main()
