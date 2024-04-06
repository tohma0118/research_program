"""
deviation(ずれ)と、外乱の終了点を確認するためのcreate_analyze_info_check.xlsxを作成する
作成されたを参考に、summary.xlsxのanalyze_infoシートを作成する
"""

import os
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font

SKIP_ROWS_NUMBER = 54
DISPLAY_DATA_RANGE = 280

file_counter = 1


def get_setting_file_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
    setting_file = os.path.join(parent_dir, "analyze_setting.txt")

    return setting_file


def get_data_dir_path(data_directory_name):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    grandparent_dir = os.path.abspath(os.path.join(current_dir, os.pardir, os.pardir))
    data_dir = os.path.join(grandparent_dir, "data", data_directory_name)

    return data_dir


def number_to_alphabet(num):
    if 7 <= num <= 16:
        return chr(ord("H") + num - 7)
    else:
        raise ValueError("Input number should be between 7 and 16")


def read_dir_name_from_settings(file_name):
    with open(file_name, "r", encoding="utf-8") as file:
        lines = file.readlines()
        for i, line in enumerate(lines):
            line = line.strip()
            if line == "directory_name" and i + 1 < len(lines):
                return lines[i + 1].strip()
    return None


def create_graph_from_csv(csv_path, ch_col, wb, i):
    global file_counter

    df = pd.read_csv(
        csv_path, encoding="Shift-JIS", skiprows=range(0, SKIP_ROWS_NUMBER)
    )

    start_col = ord(ch_col) - ord("A")
    y_values = df.iloc[:DISPLAY_DATA_RANGE, start_col].values

    sheet_title = os.path.basename(csv_path).split(".")[0]

    if "_Deoxy" in sheet_title:
        sheet_title = "Deoxy" + f"_{file_counter}" + f"_CH{i}"
    elif "_Oxy" in sheet_title:
        sheet_title = "Oxy" + f"_{file_counter}" + f"_CH{i}"
    elif "_Total" in sheet_title:
        sheet_title = "Total" + f"_{file_counter}" + f"_CH{i}"
        file_counter += 1

    ws = wb.create_sheet(title=sheet_title)

    font = Font(name="游ゴシック")

    for idx, value in enumerate(y_values, 1):
        cell = ws.cell(row=idx, column=1, value=value)
        cell.font = font

    chart = LineChart()

    # グラフのサイズを指定（ここでは横長のサイズを設定）
    chart.width = 40
    chart.height = 10

    data = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=len(y_values))
    chart.legend = None
    chart.add_data(data)
    ws.add_chart(chart, "E5")


def main():
    setting_file = get_setting_file_path()
    data_dir_name = read_dir_name_from_settings(setting_file)

    data_dir_path = get_data_dir_path(data_dir_name)

    csv_directory = os.path.join(data_dir_path, "blood_csv")
    excel_directory = os.path.join(data_dir_path, "blood_excel")

    if not os.path.isdir(excel_directory):
        os.mkdir(excel_directory)
        print("blood_excelディレクトリを作成しました。")

    output_excel_path = os.path.join(excel_directory, "analyze_info_check.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除

    # 指定したディレクトリ内のCSVファイルすべてを処理
    for i in range(7, 17):
        global file_counter
        file_counter = 1
        ch_col = number_to_alphabet(int(i))
        for csv_file in os.listdir(csv_directory):
            if csv_file.endswith(".csv"):
                create_graph_from_csv(
                    os.path.join(csv_directory, csv_file), ch_col, wb, i
                )

    wb.save(output_excel_path)
    print("エクセルファイルにグラフを保存しました。")


if __name__ == "__main__":
    main()
